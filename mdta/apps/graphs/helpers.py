from datetime import datetime
# from itertools import izip, takewhile
# from django.conf import settings
# from django.utils import timezone
from openpyxl import load_workbook
from django.db import transaction
import pandas as pd

from mdta.apps.projects.models import Project, Module, VUID


PAGE_NAME = "page name"
PROMPT_NAME = "prompt name"
PROMPT_TEXT = "prompt text"
# LANGUAGE = "language"
STATE_NAME = "state name"
DATE_CHANGED = "date changed"

VUID_HEADER_NAME_SET = {
    PROMPT_NAME,
    PROMPT_TEXT,
    DATE_CHANGED
}

@transaction.atomic
def parse_out_module_names(vuid):
    # wb = load_workbook(vuid.file.path)
    df = pd.read_excel(vuid.file.path)
    # ws = wb.active

    # df_dup = df.groupby(axis=1, level=0).apply(lambda x: x.duplicated())
    # # headers = [str(i.value).lower() for i in ws.rows[0]]
    #
    # df[~df_dup].iloc[:, 0].to_csv("media/file_unique_col1.csv")

    things = df['Page Name'].unique()
    for thing in things:
        print(thing)

    # df[~df_dup].to_csv("media/file_unique_valuesALL.csv")
    #
    # names = df['Page Name'].unique()
    # for name in names:
    #     print(df[df['Page Name'] == name])

    # try:
    #     prompt_name_i = headers.index(PROMPT_NAME)
    #     prompt_text_i = headers.index(PROMPT_TEXT)
    #     date_changed_i = headers.index(DATE_CHANGED)
    #     page_name_i = headers.index(PAGE_NAME)
    #     state_name_i = headers.index(STATE_NAME)
    # except ValueError:
    #     return {"valid": False, "message": "Parser error, invalid headers"}
    #
    # for w in ws.rows[1:]:
    #      name = str(w[prompt_name_i].value.strip())
    #      page = str(w[page_name_i].value.strip())
    #      state = str(w[state_name_i].value.strip())
    #      verbiage = str(w[prompt_text_i].value).strip()
    #      vuid_time = w[date_changed_i].value.date() if w[date_changed_i].value is datetime else None
    # print(verbiage)
    # print(tuple(ws.columns))
    # print(page)
    # print(state)
    # print(headers)
    # pandas below
    # print(df[~df_dup])
    return {"valid": True, "message": "Parsed file successfully"}


def upload_vuid(uploaded_file, user, project_id):
    vuid = VUID(filename=uploaded_file.name, file=uploaded_file, project_id=project_id, upload_by=user)
    vuid.save()

    # p = Project.objects.get(project=project.name)

    result = parse_out_module_names(vuid)
    if not result['valid']:
        vuid.delete()
        return result

    return {"valid": True, "message": "File uploaded and parsed successfully"}


# def create_modules(vuid):
#     wb = load_workbook(vuid.file.path)
#     ws = wb.active
#     indexes = [ws.columns[0].index(x) for x in set(ws.columns[0])]
#     # uq = set(ws.columns[0])
#
#     for index in indexes:
#         yield (index, wb[index])
#
#     # for cell0bj in uq:
#     #     print(cell0bj.value)



